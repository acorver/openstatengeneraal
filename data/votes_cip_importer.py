import db

import os
from copy import copy
from datetime import datetime
from openpyxl import load_workbook

# ===================================================================
# Import files received from the Centraal Informatie Punt (CIP, 
# Lit: "Central Information Point") containing voting information
# into the database format
# ===================================================================

def loadExcelTable(f):
    wb = load_workbook(filename = f)
    ws = wb.get_active_sheet() 
    for mc in ws._merged_cells:
        ws.unmerge_cells(mc)

    headers = [ws.cell(row=6, column=i).internal_value for i in xrange(0,ws.get_highest_column())]
    
    rows = []
    for r in xrange(7, ws.get_highest_row()):
        row = {}
        for c in xrange(ws.get_highest_column()):
            row[headers[c]] = ws.cell(row=r, column=c).internal_value
        rows.append(row)
    return rows

if __name__ == "__main__":
    # Process all .xlsx files provided by the CIP
    base = "./rawdata/cip/"
    for file in os.listdir(base):
        if file.endswith(".xlsx"):
            rows = loadExcelTable(base+file)
            for r in xrange(len(rows)):
                if len(rows[r]['Citeertitel'])==0:
                    continue
                # ====================
                # Parse document info
                # ====================
                d = {}
                d["date"] = str(datetime.strptime(rows[r]['Datum document'], "%Y.%m.%d"))
                d["number"] = rows[r]['Kamerstuknummer']
                d["title"] = rows[r]['Citeertitel']
                d["decision_txt"] = rows[r]['Stemming'].lower()
                
                number = d['number'].split(' ')
                if len(number) == 3:
                    d['index'] = str(int(number[0])) + '-' + number[2] + '-' + str(int(number[1]))
                elif len(number) == 2:
                    d['index'] = str(int(number[0])) + '-' + str(int(number[1]))
                else:
                    continue

                # ====================
                # Parse voting info by parties
                # and individual politicians
                # ====================
                dec = []
                try:
                    curr = r
                    for b in ["voor","tegen","verdeeld", "afwezig"]:
                        if b in rows[curr]['Stemverdeling']:
                            dec += [(x.strip(), b) for x in rows[curr]['Stemverdeling'][len(b):].split(",") if len(x)>0]
                            curr += 1
                except:
                    # Sometimes no votes are included, presumely it was accepted or rejected by an overwhelming majority
                    continue

                d["votes"] = {}
                for x in dec:
                    actorname = x[0]
                    vote = ('Y' if x[1]=='voor' else 
                            ('N' if x[1]=='tegen' else 
                             ('A' if x[1]=='afwezig' else 
                              ('divided' if x[1]=='verdeeld' else 'ERROR'))))
                    d['votes'][actorname] = vote

                db.store({"index": d["index"], "type": "vote", "cip": d})
